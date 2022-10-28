#!/bin/python
from asyncore import write
import sys,os
import time
import json

#Dependancies LIB
import openpyxl as lib
from openpyxl.worksheet.table import Table, TableStyleInfo

def initDataModel():
    return {"hostname":None,"decription":None,"deviceType":None,"IP":None,"Name":None,"OS":None}

def mapDataAttribute(index):
    maps={"0":"hostname","1":"decription","2":"IP","3":"Name"}
    if str(index) in maps:
         return maps[str(index)]
    else:
        return None

def parseData(index,data):
    if index == 2:
        return {"ip":data.split('/')[0],"mask":"255.255.255.0"}
    elif index == 101:
        data=data.upper()
        if "SWITCH" in data:
            return {"os":"ios",'type':"switch"}
        elif "ROUTER" in data:
            if "CBG" in data:
                return {"os":"ios",'type':"router"}
            else:
                return {"os":"eos",'type':"router"}
        else:
            return {"os":'other','type':"other"}
    else:
        return data

def readData(data):
    print("reading %s devices" %len(data))
    for data_value in data:
        for attr, value in data_value.items():
            print("%s : %s" %(attr,value))
        print('-'*25)

def main():
    data=[initDataModel()]
    wb = lib.load_workbook("../assets/RAWASSETS.xlsx",read_only=False)
    sheet = wb.worksheets[0]
    x= sheet.max_column
    y=sheet.max_row
    for rowList in sheet.iter_rows(): #loop through row elements
        for xIndex in range(0,x): #loop every colomn in the row
            # print(" %s %s" %(rowList[xIndex].internal_value,xIndex))
            # set constants
            constant=parseData(101,rowList[1].internal_value)
            data[len(data)-1]["OS"]=constant['os']
            data[len(data)-1]["deviceType"]=constant['type']
            attrib=mapDataAttribute(xIndex)
            if attrib:
                if rowList[xIndex].internal_value:
                    data[len(data)-1][attrib]=parseData(xIndex,rowList[xIndex].internal_value)
                else:
                    data[len(data)-1][attrib]=None
        #done with row, create append a new dict to data
        data.append(initDataModel()) 
    data.pop(0) #remove first row
    data.pop() #remove last item row
    with open('../assets/INVENTORYJSON.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    readData(data)
    return

if __name__ == '__main__':
    #sys.argv = ["programName.py","--input","test.txt","--output","tmp/test.txt"]
    main()