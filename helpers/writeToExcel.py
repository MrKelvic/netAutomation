#192.168.122.94
# from openpyxl.formatting import Rule
# from openpyxl.styles import Font, PatternFill, Border
# from openpyxl.styles.differential import DifferentialStyle
import uuid
import helpers.files as files 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill,Alignment

#Globals
headerStyle = NamedStyle(name="headerStyle")
headerStyle.font = Font(bold=True, size=11)
# bd = Side(style='medium', color="000000")
# headerStyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)

#Regulars
regularStyle = NamedStyle(name="regularStyle")
regularStyle.font = Font(bold=False, size=11)
# bd = Side(style='thin', color="000000")
# regularStyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)

def preproccess(value,key=None):
    if key not in ["state","connected","active"]:
        return value
    if type(value).__name__=="str":
        return value
    elif type(value).__name__=="bool":
        if value == True:
            return "Up"
        elif value== False:
            return "Down"
    else:
        return value

def fontGen(value,key=None):
    if type(value).__name__=="str":
        return Font(bold=False, size=11)
    elif type(value).__name__=="bool":
        if key not in ["state","connected","active"]:
            return Font(bold=False, size=11) 
        if value == True:
            return Font(bold=False, size=11,color="0000FF00")
        elif value== False:
            return Font(bold=False, size=11,color="00FF0000")
    else:
        return Font(bold=False, size=11) 

def index(output):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    
    cell=None
    tY=1
    xPadding=1
    for data in output:
        # print(data.cli)
        # print(data.node["hostname"])
        clone=data.cli
        if len(clone) <1:
            return 0
        headers=clone.pop(0)
        cell=ws1.cell(column=xPadding,row=tY,value=data.node["hostname"])
        cell.style=headerStyle
        cell.fill=PatternFill(start_color='00969696',end_color='00969696',fill_type = "solid")
        cell.alignment=Alignment(horizontal="center",vertical="center")
        tY=tY+1
        for x1,header in enumerate(headers):
            cell=ws1.cell(column=x1+xPadding,row=tY,value=str(header))
            cell.style=headerStyle
            cell.fill=PatternFill(start_color='00969696',end_color='00969696',fill_type = "solid")
            cell.alignment=Alignment(horizontal="center",vertical="center")
        tY=tY+1
        for dataArray in clone:
            for x2,cellData in enumerate(dataArray):
                cell=ws1.cell(column=x2+xPadding,row=tY,value=str(preproccess(cellData)))
                cell.style=regularStyle
                cell.font=fontGen(cellData)
            tY=tY+1
        # add extra space after each device data
        tY=tY+4
        # for x,i in enumerate(range(0,4)):
        #     ws1.cell(column=1,row=tY,value='******')
    fileName=None#input("Enter filename: ")
    if not fileName:
        fileName='Test'

    wb.save(fileName+'.xlsx')

def getReport(payload):
    REPORT_PATH="/store/temp/reports/"
    path="."+REPORT_PATH
    fileName=uuid.uuid4().hex
    files.writeToTemp(None,name="",tempdir=REPORT_PATH,extension=".",MaxDays=0)
    # 
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Automated Report"
    xPadding=1
    tY=1

    # IF isMultiple
    if payload['isMultiple']:
        # VALUES
        tYControl=4
        for nodeData in payload['values']:
            cell=ws1.cell(column=xPadding,row=tY,value=nodeData["head"])
            cell.style=headerStyle
            tY=tY+1
            # NODE HEADER
            for inc,value in enumerate(nodeData["headers"]):
                cell=ws1.cell(column=xPadding+inc,row=tY,value=value)
                cell.style=regularStyle
            tY=tY+1
            #NODE VALUES
            for iOuter,yValues in enumerate(nodeData["values"]):
                tempTy=tY
                if len(yValues) > tYControl:
                    tYControl=len(yValues)
                for iInner,yValue in enumerate(yValues):
                    cell=ws1.cell(column=xPadding+iOuter,row=tempTy+iInner,value=str(preproccess(yValue,nodeData["headers"][iOuter])))
                    cell.style=regularStyle
                    cell.font=fontGen(yValue,nodeData["headers"][iOuter])
            tY+=tYControl+3
        tY+=2
    else:
        for ia,header in  enumerate(payload['headers']):
            cell=ws1.cell(column=xPadding+ia,row=tY,value=header)
            cell.style=headerStyle
        tY=tY+1
        for keyTrack,values in enumerate(payload['values']):
            for ib,value in enumerate(values):
                cell=ws1.cell(column=xPadding+ib,row=tY,value=str(preproccess(value,payload['headers'][keyTrack+3]))) #3 is an offset for ["Device", "type", "IP"] already from the FE
                cell.style=regularStyle
                cell.font=fontGen(value,payload['headers'][keyTrack+3])
            tY=tY+1

    print("fileName ",fileName+'.xlsx')
    wb.save(path+fileName+'.xlsx')
    return REPORT_PATH[1:]+fileName+'.xlsx'

    


# if __name__ == '__main__':
#     #sys.argv = ["programName.py","--input","test.txt","--output","tmp/test.txt"]
#     index(True)